"""
Scraper utilities for the ATK List Maker page.

Architecture:
  1. Jina.ai Reader API (r.jina.ai) — free, handles JS/React pages, infinite scroll.
  2. Gemini 2.0 Flash (FREE via Google AI Studio) — extracts structured data from text.
  3. Detail page enrichment — concurrent fetching (15 workers) of individual company pages.
  4. Jina.ai Search (s.jina.ai) — fallback website URL enrichment.
"""

import re
import json
import httpx
import pandas as pd
from urllib.parse import urljoin

# ── Page fetching ─────────────────────────────────────────────────────────────

def fetch_via_jina(url: str, timeout: int = 40) -> tuple:
    """Fetch clean text content via Jina.ai Reader (free, handles JS)."""
    try:
        jina_url = f"https://r.jina.ai/{url}"
        headers = {
            "Accept": "text/plain",
            "X-No-Cache": "true",
            "X-Return-Format": "markdown",
        }
        r = httpx.get(jina_url, headers=headers, timeout=timeout, follow_redirects=True)
        if r.status_code == 200 and r.text.strip():
            return r.text, ""
        return "", f"Jina.ai returned HTTP {r.status_code}"
    except Exception as e:
        return "", f"Fetch error: {e}"


def fetch_page(url: str) -> tuple:
    """Primary fetch: tries Jina.ai first, falls back to direct request."""
    content, err = fetch_via_jina(url)
    if content:
        return content, ""
    try:
        r = httpx.get(
            url, timeout=20, follow_redirects=True,
            headers={"User-Agent": "Mozilla/5.0 (compatible; ATK-ListMaker/1.0)"},
        )
        return r.text, ""
    except Exception as e:
        return "", str(e)


# ── AI client — auto-detects OpenAI vs Gemini from key format ─────────────────

def _ai_generate(api_key: str, system: str, prompt: str) -> str:
    """
    Auto-detects provider from the API key:
      - OpenAI keys start with 'sk-'  → uses gpt-4o-mini
      - Everything else               → uses Gemini 2.0 Flash (google-genai)
    """
    if api_key.startswith("sk-"):
        # ── OpenAI ──────────────────────────────────────────────────────────
        from openai import OpenAI
        client = OpenAI(api_key=api_key)
        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": system or "You are a helpful assistant."},
                {"role": "user",   "content": prompt},
            ],
            temperature=0,
        )
        return resp.choices[0].message.content
    else:
        # ── Gemini ──────────────────────────────────────────────────────────
        from google import genai
        from google.genai import types
        client = genai.Client(api_key=api_key)
        resp = client.models.generate_content(
            model="gemini-2.0-flash",
            contents=prompt,
            config=types.GenerateContentConfig(
                system_instruction=system or "You are a helpful assistant.",
            ),
        )
        return resp.text


# ── Smart pagination ──────────────────────────────────────────────────────────

_NEXT_PAGE_SYSTEM = "You are a web scraping assistant. Analyse webpage content and return only what is asked — no explanation."

_NEXT_PAGE_PROMPT = """Current page URL: {current_url}
Base / start URL: {base_url}

Look at the page content below and find the URL of the NEXT page of results.
Look for: "Next", ">", "›", pagination numbers, or any link labelled with the next page number.

Rules:
- If you find a next-page URL, return it as a single line with no other text.
- If it is a relative URL (e.g. /exhibitors?page=2), prepend the domain to make it absolute.
- If there is NO next page (last page or only one page), return exactly: NO_MORE_PAGES
- Return NOTHING else — just the URL or NO_MORE_PAGES.

PAGE CONTENT (first 30 000 chars):
{content}"""


def find_next_page_url(content: str, current_url: str, base_url: str, api_key: str) -> str:
    """Ask Gemini to find the real next-page URL. Returns '' if no more pages."""
    try:
        prompt = _NEXT_PAGE_PROMPT.format(
            current_url=current_url,
            base_url=base_url,
            content=content[:30_000],
        )
        result = _ai_generate(api_key, _NEXT_PAGE_SYSTEM, prompt).strip()
        if not result or "NO_MORE_PAGES" in result.upper():
            return ""
        if result.startswith("http"):
            return result
        return urljoin(base_url, result)
    except Exception:
        return ""


# ── Listing page extraction ───────────────────────────────────────────────────

_EXTRACTION_SYSTEM = (
    "You are a precise data extraction assistant. "
    "Extract exhibitor data from exhibition/trade-show webpage content. "
    "Return ONLY a valid JSON array — no explanation, no markdown fences, no extra text."
)

_EXTRACTION_PROMPT = """URL scraped: {url}

User notes about this page:
{instructions}

Extract EVERY exhibitor/company listed on this page. For each one, extract:
- company_name  (required — use "" only if truly absent)
- country       (country of origin shown on this listing page, or "")
- email         (email address if shown on this listing page, or "")
- phone         (phone number if shown on this listing page, or "")
- website       (company website URL if shown on this listing page, or "")
- stand_number  (booth/stand number, or "")
- hall          (hall, pavilion, zone, or "")
- detail_url    (FULL absolute URL to this company's individual profile/detail page on this website, or "")

Critical rules:
- Include EVERY company found — do not skip any, even if they have no contact info.
- For detail_url: if the company name or card is a clickable link, extract the href as a full absolute URL.
  If the href is relative (starts with /), prepend the domain from the current page URL to make it absolute.
- Do not summarise or say "and X more" — list every single one.
- If you see a "Next" / pagination link, add one final object: {{"_meta": "has_more_pages"}}
- Return ONLY the JSON array. No markdown, no commentary.

PAGE CONTENT:
{content}"""


def extract_exhibitors(content: str, url: str, instructions: str, api_key: str) -> tuple:
    """
    Use Gemini to extract all exhibitors from page text.
    Returns (list_of_dicts, error_str, has_more_pages_hint).
    """
    try:
        prompt = _EXTRACTION_PROMPT.format(
            url=url,
            instructions=instructions.strip() if instructions else "None provided.",
            content=content[:300_000],   # 300K chars ≈ 75K tokens; fits both GPT-4o-mini and Gemini
        )
        raw = _ai_generate(api_key, _EXTRACTION_SYSTEM, prompt).strip()

        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)

        match = re.search(r"\[.*\]", raw, re.DOTALL)
        if not match:
            return [], f"Unexpected response format: {raw[:300]}", False

        data = json.loads(match.group())

        has_more = False
        clean = []
        for row in data:
            if "_meta" in row:
                if "has_more" in str(row.get("_meta", "")):
                    has_more = True
            else:
                clean.append(row)
        return clean, "", has_more

    except Exception as e:
        return [], str(e), False


# ── Detail page enrichment (concurrent) ──────────────────────────────────────

_SKIP_DOMAINS = {
    "facebook.com", "twitter.com", "linkedin.com", "instagram.com",
    "youtube.com", "wikipedia.org", "google.com", "jina.ai",
    "bloomberg.com", "crunchbase.com", "trustpilot.com",
}

_EMAIL_RE = re.compile(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}')
_PHONE_RE = re.compile(
    r'(?:\+\d{1,3}[\s\-]?)?\(?\d{2,4}\)?[\s\-]?\d{3,5}[\s\-]?\d{3,6}(?:[\s\-]?\d{2,4})?'
)
_URL_RE = re.compile(r'https?://[^\s\)\"\'\]<>,;]+')


def _regex_extract_contact(content: str, extra_skip: set = None) -> dict:
    """Fast regex extraction of email, phone, website from Jina markdown."""
    skip = _SKIP_DOMAINS | (extra_skip or set())
    result = {"email": "", "phone": "", "website": ""}

    emails = _EMAIL_RE.findall(content)
    if emails:
        result["email"] = emails[0]

    phones = _PHONE_RE.findall(content)
    if phones:
        result["phone"] = phones[0].strip()

    for url in _URL_RE.findall(content):
        dm = re.search(r'https?://(?:www\.)?([^/]+)', url)
        if dm:
            d = dm.group(1).lower()
            if not any(s in d for s in skip):
                result["website"] = url.rstrip(".,;)")
                break

    return result


def enrich_detail_pages(rows: list, api_key: str, site_domain: str = "",
                        progress_callback=None) -> list:
    """
    For rows that have a detail_url, fetch each company's profile page
    and fill in missing email, phone, website, country.

    Uses 15 concurrent threads — ~1 min per 200 companies.
    progress_callback(done, total, company_name) called after each company.
    Returns updated list of row dicts.
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    # Skip the exhibitor directory's own domain when looking for company websites
    extra_skip = set()
    if site_domain:
        extra_skip.add(site_domain.lstrip("www.").lower())

    needs_detail = [
        (i, row) for i, row in enumerate(rows)
        if row.get("detail_url", "").strip().startswith("http")
    ]

    if not needs_detail:
        return rows

    results = [dict(r) for r in rows]

    def _process(idx_row):
        idx, row = idx_row
        detail_url = row["detail_url"].strip()
        content, _ = fetch_via_jina(detail_url, timeout=30)
        if not content:
            return idx, {}

        # Fast regex first — uses no Gemini quota
        contact = _regex_extract_contact(content, extra_skip=extra_skip)

        # Fall back to Gemini only if regex found nothing useful
        if not contact["email"] and not contact["website"] and api_key:
            try:
                prompt = (
                    "Extract contact info from this company profile page. "
                    "Return ONLY a JSON object with keys: email, phone, website, country. "
                    "Use empty string for anything not found. "
                    "For 'website': return the company's OWN website, not the exhibition directory URL.\n\n"
                    f"PAGE CONTENT:\n{content[:15_000]}"
                )
                raw = _ai_generate(
                    api_key,
                    "You extract contact info. Return only a JSON object.",
                    prompt,
                ).strip()
                raw = re.sub(r"^```(?:json)?\s*", "", raw)
                raw = re.sub(r"\s*```$", "", raw)
                m = re.search(r"\{.*\}", raw, re.DOTALL)
                if m:
                    contact = json.loads(m.group())
            except Exception:
                pass

        return idx, contact

    total = len(needs_detail)
    done_count = [0]

    with ThreadPoolExecutor(max_workers=15) as executor:
        futures = {executor.submit(_process, item): item for item in needs_detail}
        for future in as_completed(futures):
            idx, contact = future.result()
            row = results[idx]
            for field in ("email", "phone", "website", "country"):
                if contact.get(field) and not row.get(field):
                    row[field] = contact[field]
            done_count[0] += 1
            if progress_callback:
                progress_callback(done_count[0], total, row.get("company_name", ""))

    return results


# ── Website enrichment via Jina.ai search ────────────────────────────────────

def enrich_missing_websites(rows: list, progress_callback=None) -> list:
    """
    For rows where website == "", search Jina.ai to find the company's official website.
    progress_callback(company_name) is called before each search if provided.
    """
    enriched = []
    for row in rows:
        name = row.get("company_name", "").strip()
        if not row.get("website") and name:
            if progress_callback:
                progress_callback(name)
            try:
                q = f"{name} official website"
                r = httpx.get(
                    f"https://s.jina.ai/{httpx.URL(q)}",
                    timeout=15,
                    headers={"Accept": "text/plain"},
                )
                if r.status_code == 200:
                    urls = re.findall(r'https?://[^\s\)\"\'\]]+', r.text)
                    for found in urls:
                        domain = re.search(r'https?://([^/]+)', found)
                        if domain:
                            d = domain.group(1).lstrip("www.")
                            if not any(skip in d for skip in _SKIP_DOMAINS):
                                row = dict(row)
                                row["website"] = found.rstrip(".,;)")
                                break
            except Exception:
                pass
        enriched.append(row)
    return enriched


# ── DataFrame helpers ─────────────────────────────────────────────────────────

_COL_RENAME = {
    "company_name": "Company Name",
    "country":      "Country",
    "email":        "Email",
    "phone":        "Phone",
    "website":      "Website",
    "stand_number": "Stand Number",
    "hall":         "Hall / Pavilion",
}

_COL_ORDER = ["Company Name", "Country", "Email", "Phone", "Website", "Stand Number", "Hall / Pavilion"]


def rows_to_df(rows: list) -> pd.DataFrame:
    """Convert extracted row dicts to a clean, consistently-ordered DataFrame."""
    if not rows:
        return pd.DataFrame(columns=list(_COL_RENAME.values()))
    df = pd.DataFrame(rows)
    df = df.drop(columns=["detail_url"], errors="ignore")   # internal field, not for output
    df = df.rename(columns={k: v for k, v in _COL_RENAME.items() if k in df.columns})
    ordered = [c for c in _COL_ORDER if c in df.columns]
    extra = [c for c in df.columns if c not in ordered]
    df = df[ordered + extra]
    df = df.fillna("").astype(str).replace("nan", "").replace("None", "")
    return df


def df_to_excel_bytes(df: pd.DataFrame) -> bytes:
    """Return Excel file bytes with ATK orange header styling."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Exhibitors"

    header_fill = PatternFill("solid", fgColor="FF6600")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")

    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = max(len(str(col_name)),
                      max((len(str(v)) for v in df.iloc[:, col_idx - 1]), default=0))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
